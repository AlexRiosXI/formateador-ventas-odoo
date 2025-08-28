from PyQt5.QtWidgets import QApplication, QFileDialog
import openpyxl
import sys
import re

output_file = input("Nombre del archivo de salida: ")
preserve_commas = input("Conservar comas en el nombre del producto? (s/n): ")
if preserve_commas == "s":
    preserve_commas = True
    print("se usara un | como separador de campos")
else:
    preserve_commas = False
    

app = QApplication(sys.argv)
file, _ = QFileDialog.getOpenFileName(None, "Selecciona un archivo", "Archivos de datos (*.xlsx *.csv)")
if not file:
    exit()

file = openpyxl.load_workbook(file)

print("Verificando columnas...")
columns = file.active.iter_rows(min_row=1, max_row=1, values_only=True)
target_columns = ["date","orderlinesproduct", "orderlinesproductinternalreference", "orderlinesquantity", "orderlinesdiscount", "orderlinessubtotalwotax", "orderlinessubtotal", "salesteam"]
for column in columns:
    for n, col in enumerate(column):
        if type(col) == str:
            col = re.sub(r'[^a-zA-Z]', '', col).lower()
            if n >= len(target_columns):
                break
            if target_columns[n] != col:
                raise Exception(f"Columna {n+1} es {col} pero debería ser {target_columns[n]}")

print("Columnas correctas")
                
sale_index = 0
sale_date = None
sale_branch = None
rows = []
print("Creando archivo de salida")
with open(f"{output_file}.csv", "w") as f:
    f.write("NoVenta,Fecha,Sucursal,Producto,SKU,Cantidad,Descuento,SubtotalAntesImpuestos,Subtotal\n")
print("Operando filas...")
total_rows = file.active.max_row
separator = "|" if preserve_commas else ","
for row_index, row in enumerate(file.active.iter_rows(min_row=2, values_only=True)):
    print(f"Procesando fila {row_index+1} de {total_rows}", end="\r")
    #Handle Sale ID
    if row[0]:
        sale_index += 1
        sale_date = row[0] 
        sale_branch = row[7] if row[7] else "No se incluyo la sucursal en el archivo original"
    sku = row[2] if row[2] else ""
    description = row[1] if row[1] else ""
    if sku:
        description = description.replace(f"[{sku}]", "")
    
    quantity = str(row[3]) if row[3] else ""
    discount = str(row[4]) if row[4] else ""
    subtotal_before_tax = str(row[5]) if row[5] else ""
    subtotal = str(row[6]) if row[6] else ""
    if not preserve_commas:
        sale_branch = sale_branch.replace(f",", " ")
        description = description.replace(f",", " ")
        quantity = quantity.replace(f",", ".")
        discount = discount.replace(f",", ".")
        subtotal_before_tax = subtotal_before_tax.replace(f",", ".")
        subtotal = subtotal.replace(f",", ".")
    
    
    with open(f"{output_file}.csv", "a") as f:
        f.write(f"{sale_index}{separator}{sale_date}{separator}{sale_branch}{separator}{description}{separator}{sku}{separator}{quantity}{separator}{discount}{separator}{subtotal_before_tax}{separator}{subtotal}\n")

print(f"Archivo terminado con exito, {output_file}.csv")
    
